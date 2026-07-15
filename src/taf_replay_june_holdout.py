"""Locked June 2026 rain holdout for the frozen WAWP TAF candidates.

The June workbook has separate ``bank data taf`` and ``bank data metar``
sheets. It contains METAR text but no independent rainfall-amount sheet, so
rain verification uses explicit RA/TSRA weather reports only. This definition
is intentionally kept separate from the January-May rainfall-amount archive.
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import sqlite3
from collections import defaultdict
from datetime import date, datetime, time, timedelta, timezone
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from src.legacy_taf_verification import active_state_structured, parse_taf, repair_validity_to_24h
from src.taf_replay_event_window_verification import _event_scores, _matching_counts
from src.taf_replay_experiment import (
    MODELS,
    AsOfWeightEngine,
    HistoricalPrior,
    ReplayConfig,
    _build_consensus,
    _flatten_taf,
    _historical_window,
    _issue_times,
    _load_historical_forecasts,
    _load_pairs,
)
from src.taf_replay_gate_sweep import GatePolicy, _policy_config
from src.tafor_generator import generate_tafor


UTC = timezone.utc
PERIOD = "2026-06"
START = pd.Timestamp("2026-06-01")
END = pd.Timestamp("2026-06-30")
RAIN_RE = re.compile(r"(?<![A-Z0-9])(?:[-+]?RA|[-+]?SHRA|[-+]?TSRA)(?![A-Z0-9])")
TS_RE = re.compile(r"(?<![A-Z0-9])TS(?:RA|GR|GS)?(?![A-Z0-9])")
POLICIES = (
    GatePolicy("control_current", "Current operational rain gate.", 40.0, 0.0, 0.0, 2, "broad_current"),
    GatePolicy("rain_50_20", "Frozen balanced candidate from the January-May sweep.", 50.0, 0.20, 0.0, 0, "broad_current"),
    GatePolicy("rain_50_20_becmg3", "Frozen wording candidate: 50/20 occurrence gate and 3-hour, 25%-agreement BECMG qualification.", 50.0, 0.20, 0.0, 0, "broad_current", becmg_min_rain_hours=3, becmg_min_agreement=0.25),
)


def _write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    fields = list(rows[0]) if rows else ["empty"]
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def _excel_date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        return from_excel(value).date()
    raise ValueError(f"Unsupported Excel date value: {value!r}")


def _time_parts(value: Any) -> tuple[int, int]:
    if isinstance(value, datetime):
        return value.hour, value.minute
    if isinstance(value, time):
        return value.hour, value.minute
    raise ValueError(f"Unsupported Excel time value: {value!r}")


def _weather_has_rain(text: str) -> bool:
    return bool(RAIN_RE.search(str(text or "").upper()))


def _weather_has_thunderstorm(text: str) -> bool:
    return bool(TS_RE.search(str(text or "").upper()))


def extract_june_workbook(workbook_path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    """Extract raw METAR rows, hourly METAR rows, and original TAFs without edits."""
    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    try:
        taf_sheet = workbook["bank data taf"]
        metar_sheet = workbook["bank data metar"]
        tafs: list[dict[str, Any]] = []
        current_date: date | None = None
        for row_number, values in enumerate(taf_sheet.iter_rows(min_row=2, values_only=True), start=2):
            if values[0] is not None:
                current_date = _excel_date(values[0])
            taf_text = str(values[2] or "").strip()
            if not taf_text or "TAF WAWP" not in taf_text:
                continue
            source = parse_taf(taf_text, PERIOD)
            parsed, repaired, reported_hours = repair_validity_to_24h(source)
            tafs.append({
                "configuration": "human_original",
                "source_taf_row": row_number,
                "source_date_anchor": current_date.isoformat() if current_date else "",
                "source_time": str(values[1] or ""),
                "issuance_utc": parsed.issue_time.isoformat().replace("+00:00", "Z"),
                "valid_start_utc": parsed.valid_start.isoformat().replace("+00:00", "Z"),
                "valid_end_utc": parsed.valid_end.isoformat().replace("+00:00", "Z"),
                "reported_validity_hours": reported_hours,
                "validity_repaired_to_24h": repaired,
                "taf": parsed.text,
            })

        raw_metar: list[dict[str, Any]] = []
        hourly_metar: list[dict[str, Any]] = []
        current_date = None
        for row_number, values in enumerate(metar_sheet.iter_rows(min_row=2, values_only=True), start=2):
            if values[1] is not None:
                current_date = _excel_date(values[1])
            text = str(values[4] or "").strip()
            if current_date is None or not text.startswith("METAR WAWP"):
                continue
            hour, minute = _time_parts(values[3])
            observed_at = datetime.combine(current_date, time(hour, minute), tzinfo=UTC)
            row = {
                "source_metar_row": row_number,
                "observed_at_utc": observed_at.isoformat().replace("+00:00", "Z"),
                "metar_day_utc": observed_at.day,
                "metar_hour_utc": hour,
                "metar_minute_utc": minute,
                "metar_text": text,
                "observed_rain_from_metar_text": _weather_has_rain(text),
                "observed_thunderstorm_from_metar_text": _weather_has_thunderstorm(text),
            }
            raw_metar.append(row)
            if minute == 0:
                hourly_metar.append(row)
        return raw_metar, hourly_metar, tafs
    finally:
        workbook.close()


def _replay_tafs(
    db_path: Path, human_tafs: list[dict[str, Any]]
) -> tuple[list[dict[str, Any]], list[str], list[str]]:
    human_issues = {
        pd.Timestamp(row["issuance_utc"].replace("Z", "+00:00")).tz_localize(None)
        for row in human_tafs
    }
    official_issues = set(_issue_times(START, END))
    nonstandard = sorted(
        timestamp.strftime("%Y-%m-%d %H:%M:%S")
        for timestamp in human_issues
        if timestamp not in official_issues
    )
    with sqlite3.connect(db_path) as connection:
        pairs = _load_pairs(connection)
        forecasts = _load_historical_forecasts(connection, START, END + pd.Timedelta(days=1))
    prior = HistoricalPrior(pairs, START.normalize())
    weight_engine = AsOfWeightEngine(pairs)
    base_config = ReplayConfig("raw_asof60", "asof", 60, "none")
    policy_configs = {policy.name: _policy_config(policy) for policy in POLICIES}
    weights_cache: dict[pd.Timestamp, dict[str, dict[str, float]]] = {}
    output: list[dict[str, Any]] = []
    skipped: list[str] = []
    for issue_utc in _issue_times(START, END):
        if issue_utc not in human_issues:
            continue
        rows = _historical_window(forecasts, issue_utc)
        if len(rows) < 24 * len(MODELS) * 0.55:
            skipped.append(issue_utc.strftime("%Y-%m-%d %H:%M:%S"))
            continue
        asof_cutoff = issue_utc + pd.Timedelta(hours=8)
        if asof_cutoff not in weights_cache:
            weights_cache[asof_cutoff] = weight_engine.weights(asof_cutoff, lookback_days=60, equal=False)
        weights = weights_cache[asof_cutoff]
        consensus, model_data, qm_rain = _build_consensus(rows, issue_utc, weights, prior, base_config)
        for policy in POLICIES:
            taf = generate_tafor(consensus, model_data, qm_rain, weights, experimental_taf_config=policy_configs[policy.name])
            if taf and taf.get("taf_text"):
                output.append({
                    "configuration": policy.name,
                    "issuance_utc": issue_utc.strftime("%Y-%m-%dT%H:%M:%SZ"),
                    "taf": _flatten_taf(taf["taf_text"]),
                    "policy": json.dumps(policy.__dict__, sort_keys=True),
                })
    return output, skipped, nonstandard


def _event_rows(configuration: str, taf_text: str, issuance_utc: str, observations: dict[datetime, dict[str, Any]]) -> list[dict[str, Any]]:
    source = parse_taf(taf_text, PERIOD)
    taf, _, _ = repair_validity_to_24h(source)
    rows: list[dict[str, Any]] = []
    current = taf.valid_start
    while current < taf.valid_end:
        observation = observations.get(current)
        if observation:
            _, state, _ = active_state_structured(taf, current)
            rows.append({
                "configuration": configuration,
                "issuance_utc": issuance_utc,
                "issuance_key": f"{configuration}|{issuance_utc}",
                "valid_time": current,
                "forecast_events": {
                    "rain_any": _weather_has_rain(state.weather or ""),
                    "thunderstorm": _weather_has_thunderstorm(state.weather or ""),
                },
                "observed_events": {
                    "rain_any": bool(observation["observed_rain_from_metar_text"]),
                    "thunderstorm": bool(observation["observed_thunderstorm_from_metar_text"]),
                },
                "forecast_weather": state.weather or "",
                "metar_text": observation["metar_text"],
            })
        current += timedelta(hours=1)
    return rows


def _all_hour_accuracy(rows: list[dict[str, Any]], event: str) -> dict[str, Any]:
    """Exact-hour binary accuracy, matching the workbook's dry/wet equality idea."""
    hits = misses = false_alarms = correct_negatives = 0
    for row in rows:
        forecast = bool(row["forecast_events"][event])
        observed = bool(row["observed_events"][event])
        if forecast and observed:
            hits += 1
        elif forecast:
            false_alarms += 1
        elif observed:
            misses += 1
        else:
            correct_negatives += 1
    total = len(rows)
    accuracy = (hits + correct_negatives) / total if total else None
    specificity = correct_negatives / (correct_negatives + false_alarms) if correct_negatives + false_alarms else None
    balanced_accuracy = (
        ((hits / (hits + misses)) + specificity) / 2
        if hits + misses and specificity is not None
        else None
    )
    return {
        "hourly_hits": hits,
        "hourly_misses": misses,
        "hourly_false_alarms": false_alarms,
        "hourly_correct_negatives": correct_negatives,
        "all_hour_accuracy": round(accuracy, 4) if accuracy is not None else None,
        "specificity": round(specificity, 4) if specificity is not None else None,
        "balanced_accuracy": round(balanced_accuracy, 4) if balanced_accuracy is not None else None,
    }


def _summarize_events(rows: list[dict[str, Any]]) -> dict[str, Any]:
    output: dict[str, Any] = {}
    by_config: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        by_config[row["configuration"]].append(row)
    for configuration, config_rows in sorted(by_config.items()):
        per_event: dict[str, Any] = {}
        for event in ("rain_any", "thunderstorm"):
            totals = defaultdict(int)
            by_issuance: dict[str, list[dict[str, Any]]] = defaultdict(list)
            for row in config_rows:
                by_issuance[row["issuance_key"]].append(row)
            for issuance_rows in by_issuance.values():
                counts, _ = _matching_counts(issuance_rows, event, window_hours=2)
                for key, value in counts.items():
                    totals[key] += value
            per_event[event] = {
                **_event_scores(dict(totals), []),
                **_all_hour_accuracy(config_rows, event),
            }
        output[configuration] = {
            "issuance_count": len({row["issuance_utc"] for row in config_rows}),
            "hourly_observation_count": len(config_rows),
            "events": per_event,
        }
    return output


def _write_report(path: Path, summary: dict[str, Any]) -> None:
    lines = [
        "# WAWP Locked June 2026 Rain Holdout",
        "",
        "## Scope",
        "",
        "This is an untouched June comparison of the frozen human TAFs and three previously selected machine configurations. The source workbook uses `bank data taf` and `bank data metar`. It is not modified.",
        "",
        "Rain verification uses explicit RA/TSRA weather wording in the hourly METAR text because this workbook has no separate rainfall-amount source. It must not be pooled directly with the January-May amount-based rain score.",
        "",
        "## Coverage",
        "",
        f"- Original human TAFs extracted: {summary['human_taf_count']}",
        f"- Human TAFs on official issuance slots: {summary['official_human_taf_count']}",
        f"- Shared human/machine TAF issuances scored: {summary['common_issuance_count']}",
        f"- Half-hourly METARs extracted: {summary['raw_metar_count']}",
        f"- Hourly METARs used for scoring: {summary['hourly_metar_count']}",
        f"- Archived forecast issuances unavailable for full replay: {len(summary['skipped_replay_issuances'])}",
        f"- Nonstandard human issuance(s) excluded from official comparison: {len(summary['nonstandard_human_issuances'])}",
        "",
        "## Event-Window Scores",
        "",
        "One-to-one matching is within plus/minus two hours inside each TAF validity. Lower FAR and higher CSI are better. All-hour accuracy is the exact-hour dry/wet agreement rate, analogous to the workbook Endapan score.",
        "",
        "| Configuration | Rain all-hour accuracy | Rain POD | Rain FAR | Rain CSI | Rain hits | Rain false alarms | TS POD | TS FAR | TS CSI |",
        "| --- | ---: | ---: | ---: | ---: | ---: | ---: | ---: | ---: | ---: |",
    ]
    for configuration, values in summary["scores"].items():
        rain = values["events"]["rain_any"]
        ts = values["events"]["thunderstorm"]
        def pct(value: Any) -> str:
            return "n/a" if value is None else f"{float(value) * 100:.1f}%"
        lines.append(
            f"| `{configuration}` | {pct(rain['all_hour_accuracy'])} | {pct(rain['POD'])} | {pct(rain['FAR'])} | {pct(rain['CSI'])} | {rain['hits']} | {rain['false_alarms']} | {pct(ts['POD'])} | {pct(ts['FAR'])} | {pct(ts['CSI'])} |"
        )
    lines.extend([
        "",
        "## Guardrails",
        "",
        "- Thresholds were frozen before extracting the June score: no June result is used to retune them.",
        "- The final June boundary may have incomplete replay coverage if the historical forecast archive ends before its 24-hour validity window. Nonstandard human amendments and skipped official issuances are named in `june_holdout_summary.json`.",
        "- This is an external holdout, but it is one month and a text-weather rain definition. It informs the production decision; it does not replace live shadow verification.",
    ])
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def run(workbook_path: Path, db_path: Path, output_dir: Path) -> dict[str, Any]:
    output_dir.mkdir(parents=True, exist_ok=True)
    raw_metar, hourly_metar, human_tafs = extract_june_workbook(workbook_path)
    observations = {
        datetime.fromisoformat(row["observed_at_utc"].replace("Z", "+00:00")): row
        for row in hourly_metar
    }
    machine_tafs, skipped, nonstandard = _replay_tafs(db_path, human_tafs)
    machine_issuances = {row["issuance_utc"] for row in machine_tafs}
    evaluated_human_tafs = [row for row in human_tafs if row["issuance_utc"] in machine_issuances]
    event_rows: list[dict[str, Any]] = []
    for row in evaluated_human_tafs + machine_tafs:
        event_rows.extend(_event_rows(row["configuration"], row["taf"], row["issuance_utc"], observations))
    csv_rows = [{
        **{key: value for key, value in row.items() if key not in {"valid_time", "forecast_events", "observed_events"}},
        "valid_time_utc": row["valid_time"].isoformat().replace("+00:00", "Z"),
        "forecast_rain": row["forecast_events"]["rain_any"],
        "observed_rain": row["observed_events"]["rain_any"],
        "forecast_thunderstorm": row["forecast_events"]["thunderstorm"],
        "observed_thunderstorm": row["observed_events"]["thunderstorm"],
    } for row in event_rows]
    summary = {
        "scope": "locked June holdout; no operational changes",
        "workbook": str(workbook_path),
        "rain_observation_definition": "explicit RA/SHRA/TSRA wording in hourly METAR text",
        "human_taf_count": len(human_tafs),
        "official_human_taf_count": len(human_tafs) - len(nonstandard),
        "common_issuance_count": len({row["issuance_utc"] for row in evaluated_human_tafs}),
        "raw_metar_count": len(raw_metar),
        "hourly_metar_count": len(hourly_metar),
        "machine_taf_count": len(machine_tafs),
        "skipped_replay_issuances": skipped,
        "nonstandard_human_issuances": nonstandard,
        "scores": _summarize_events(event_rows),
    }
    _write_csv(output_dir / "june_metar_all_half_hour.csv", raw_metar)
    _write_csv(output_dir / "june_metar_hourly.csv", hourly_metar)
    _write_csv(output_dir / "june_human_tafs.csv", human_tafs)
    _write_csv(output_dir / "june_machine_tafs.csv", machine_tafs)
    _write_csv(output_dir / "june_holdout_hourly.csv", csv_rows)
    (output_dir / "june_holdout_summary.json").write_text(json.dumps(summary, indent=2) + "\n", encoding="utf-8")
    _write_report(output_dir / "JUNE_HOLDOUT_REPORT.md", summary)
    return summary


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, default=Path(r"C:\Users\MY ASUS\Downloads\Verifikasi TAF_FORM_Juni_2026 (2).xlsx"))
    parser.add_argument("--db", type=Path, default=Path(r"D:\UJI_PERFORMA_MODEL\meteologix-wawp-main\wawp_forecasts.db"))
    parser.add_argument("--output-dir", type=Path, default=Path(r"D:\UJI_PERFORMA_MODEL\VERIFICATION_REPORTS\taf_june_holdout_2026"))
    args = parser.parse_args()
    result = run(args.workbook, args.db, args.output_dir)
    print(json.dumps({"output_dir": str(args.output_dir), "scores": result["scores"]}, indent=2))


if __name__ == "__main__":
    main()
