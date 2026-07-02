# -*- coding: utf-8 -*-
"""
==================================================================
  METEOLOGIX MULTI-MODEL SCRAPER & DECODER
  Sourcing ECMWF, GFS, ICON, ACCESS-G3, UKMO, GEM, and Multi-Model
  Bandara Sangia Ni Bandera (WAWP) - Kolaka, Indonesia
==================================================================
This script runs in parallel with the Windy scraper to collect 
and archive multi-model data from Meteologix for future migration evaluation.

It outputs:
  1. Excel sheet per model in a consolidated workbook with extended parameters.
  2. A flat Parquet sidecar stacking all models (aligned with system schema).
  3. Daily and Latest JSON dumps.
==================================================================
"""

import os
import re
import sys
import json
import time
import logging
import pandas as pd
from datetime import datetime, timezone, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ================== CONFIGURATION ==================
STATION_NAME = "Bandara_Sangia_Ni_Bandera"
MODEL_URLS = {
    "ECMWF": "https://meteologix.com/sg/forecast/8299403-sangia-nibandera-airport/meteogram/euro",
    "GFS": "https://meteologix.com/sg/forecast/8299403-sangia-nibandera-airport/meteogram/usa",
    "ICON": "https://meteologix.com/sg/forecast/8299403-sangia-nibandera-airport/meteogram/deu",
    "ACCESS-G3": "https://meteologix.com/sg/forecast/8299403-sangia-nibandera-airport/meteogram/aus",
    "UKMO": "https://meteologix.com/sg/forecast/8299403-sangia-nibandera-airport/meteogram/gbr",
    "GEM": "https://meteologix.com/sg/forecast/8299403-sangia-nibandera-airport/meteogram/can",
    "Multi-Model": "https://meteologix.com/sg/forecast/8299403-sangia-nibandera-airport/meteogram/multimodel"
}

_HERE = os.path.dirname(os.path.abspath(__file__))
ARCHIVE_DIR = os.path.join(os.path.dirname(_HERE), "Archives", "Meteologix_MultiModel")
WITA = timezone(timedelta(hours=8))
KPH_PER_KNOT = 1.852

# Model update schedules are now extracted live from the page's
# <div class="model-update-info"> element during scraping.
# No hardcoded schedule assumptions needed.

# ================== LOGGING ==================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S"
)
log = logging.getLogger("meteologix_multi")

# Peta simbol cuaca -> label Indonesia
SIMBOL = {
    "sunshine": "Cerah", "sunshine_night": "Cerah (malam)",
    "partlycloudy": "Berawan sebagian", "partlycloudy_night": "Berawan sebagian (malam)",
    "cloudy": "Berawan", "overcast": "Mendung",
    "rain": "Hujan", "lightrain": "Hujan ringan", "heavyrain": "Hujan lebat",
    "thunderstorm": "Badai petir", "snow": "Salju", "fog": "Kabut",
}

# ================== SCRAPER & PARSER ==================

def parse_run_init(text: str) -> datetime | None:
    """
    Extract UTC initialization time from the Meteologix page.
    
    Tries two strategies in order:
      1. HTML div text: "Forecast from DD.MM.YYYY, HHz (UTC)"
      2. JS variable fallback: hccompact_model_starttime = <epoch_ms>
    """
    if not text:
        return None
    
    # Strategy 1: Parse from <div class="model-run-info"> text
    # Format: "Forecast from 01.07.2026, 18z (UTC)"
    m = re.search(
        r'Forecast\s+from\s+(\d{2})\.(\d{2})\.(\d{4}),?\s*(\d{1,2})z',
        text, re.IGNORECASE
    )
    if m:
        day, month, year, hour = int(m.group(1)), int(m.group(2)), int(m.group(3)), int(m.group(4))
        return datetime(year, month, day, hour, 0, 0, tzinfo=timezone.utc)
    
    # Strategy 2: Fallback to JS variable (legacy)
    m = re.search(r'hccompact_model_starttime\s*=\s*["\']?(\d+)["\']?', text)
    if m:
        ts_ms = int(m.group(1))
        return datetime.fromtimestamp(ts_ms / 1000, timezone.utc)
        
    return None

def scrape_meteogram_script(url: str) -> tuple:
    """
    Fetch the Highcharts script block AND model metadata from a Meteologix page.
    
    Returns:
        (script_text, run_info_text, update_info_text)
        - script_text:      The JS containing hccompact_data_* variables.
        - run_info_text:     Text from <div class="model-run-info">, e.g.
                            "Forecast from 01.07.2026, 18z (UTC)"
        - update_info_text:  Text from <div class="model-update-info">, e.g.
                            "Update: 4 times a day, approx. 09:00am, ..."
    """
    from selenium import webdriver
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.common.exceptions import TimeoutException
    from webdriver_manager.chrome import ChromeDriverManager

    opt = Options()
    opt.add_argument("--headless=new")
    opt.add_argument("--no-sandbox")
    opt.add_argument("--disable-dev-shm-usage")
    opt.add_argument("--window-size=1920,1080")
    opt.add_argument("--disable-gpu")
    opt.add_argument("--lang=en-US")
    opt.add_argument(
        "--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/148.0.0.0 Safari/537.36"
    )

    driver = webdriver.Chrome(
        service=Service(ChromeDriverManager().install()), options=opt
    )
    try:
        driver.get(url)
        try:
            WebDriverWait(driver, 20).until(
                lambda d: d.execute_script(
                    "var scripts = Array.from(document.querySelectorAll('script'));"
                    "return scripts.some(function(s) { "
                    "  return s.textContent && s.textContent.indexOf('hccompact_data_symbols') > -1; "
                    "});"
                )
            )
        except TimeoutException:
            log.warning("Timeout waiting for hccompact variables on: %s", url)

        # ── Extract model metadata from HTML divs ──
        run_info_text = ""
        update_info_text = ""
        try:
            el = driver.find_element(By.CSS_SELECTOR, "div.model-run-info")
            run_info_text = el.text.strip()
            log.info("  Model run info: %s", run_info_text)
        except Exception:
            log.warning("  Could not find div.model-run-info on: %s", url)
        
        try:
            el = driver.find_element(By.CSS_SELECTOR, "div.model-update-info")
            update_info_text = el.text.strip()
            log.info("  Model update schedule: %s", update_info_text)
        except Exception:
            log.debug("  Could not find div.model-update-info on: %s", url)

        # ── Extract all script tags text ──
        scripts = driver.execute_script(
            "return Array.from(document.querySelectorAll('script'))"
            ".map(function(s){return s.textContent||'';});"
        )
        for txt in scripts:
            if "hccompact_data_symbols" in txt or "hccompact_data_direction" in txt:
                return txt, run_info_text, update_info_text
        
        raise RuntimeError("No script tag containing meteogram data found.")
    finally:
        driver.quit()

def _blok_data(txt: str, name: str):
    m = re.search(
        r"name:\s*'" + re.escape(name) + r"'.*?data:\s*\[(.*?)\]\s*,\s*"
        r"(?:color|\}|[A-Za-z_]+\s*:)",
        txt, re.S,
    )
    if not m:
        m = re.search(
            r"name:\s*'" + re.escape(name) + r"'.*?data:\s*\[(.*?)\]\s*\}",
            txt, re.S,
        )
    return m.group(1) if m else None

def seri_pasangan(txt: str, name: str) -> dict:
    body = _blok_data(txt, name)
    if body is None:
        return {}
    out = {}
    for a, b in re.findall(r"\[(\d+),\s*([\d.\-]+)\]", body):
        out[int(a)] = float(b)
    return out

def seri_pasangan_pattern(txt: str, pattern: str) -> dict:
    """Find data block using a regex pattern on the series name (robust for HTML names)."""
    m = re.search(
        r"name:\s*['\"].*?" + re.escape(pattern) + r".*?['\"].*?data:\s*\[(.*?)\]\s*,\s*"
        r"(?:color|\}|[A-Za-z_]+\s*:)",
        txt, re.S,
    )
    if not m:
        m = re.search(
            r"name:\s*['\"].*?" + re.escape(pattern) + r".*?['\"].*?data:\s*\[(.*?)\]\s*\}",
            txt, re.S,
        )
    if not m:
        return {}
    body = m.group(1)
    out = {}
    for a, b in re.findall(r"\[(\d+),\s*([\d.\-]+)\]", body):
        out[int(a)] = float(b)
    return out

def seri_kolom(txt: str, name: str) -> dict:
    body = _blok_data(txt, name)
    if body is None:
        return {}
    out = {}
    for obj in re.findall(r'\{[^}]*\}', body):
        mx = re.search(r'"?x"?:\s*(\d+)', obj)
        my = re.search(r'"?y"?:\s*([\d.\-]+)', obj)
        if mx and my:
            out[int(mx.group(1))] = float(my.group(1))
    return out

def seri_array(txt: str, varname: str) -> list:
    m = re.search(re.escape(varname) + r"\s*=\s*(\[.*?\]);", txt, re.S)
    if not m:
        return []
    try:
        return json.loads(m.group(1))
    except json.JSONDecodeError:
        return re.findall(r"'([^']*)'", m.group(1))

def seri_awan(txt: str) -> dict:
    """Parse Clouds (%) range multi-layer data [timestamp, layer_idx, percentage]."""
    m = re.search(
        r"name:\s*['\"]Clouds \(%\)['\"].*?data:\s*\[(.*?)\]\s*,\s*"
        r"(?:color|\}|[A-Za-z_]+\s*:)",
        txt, re.S,
    )
    if not m:
        m = re.search(
            r"name:\s*['\"]Clouds \(%\)['\"].*?data:\s*\[(.*?)\]\s*\}",
            txt, re.S,
        )
    if not m:
        return {}
    body = m.group(1)
    out = {}
    # Match [timestamp, layer_idx, percentage]
    for ts, layer, val in re.findall(r"\[(\d+),\s*(\d+),\s*([\d.\-]+)\]", body):
        t = int(ts)
        lay = int(layer)
        v = float(val)
        if t not in out:
            out[t] = {0: 0.0, 1: 0.0, 2: 0.0}
        out[t][lay] = v
    return out

def decode_model_script(txt: str, model_name: str, run_init_utc: datetime = None) -> list:
    """Decode Highcharts script variables into a structured list of dictionaries."""
    # Standard Pairwise series
    temp = seri_pasangan_pattern(txt, 'Temperature')
    dew = seri_pasangan_pattern(txt, 'Dew point')
    gust = seri_pasangan_pattern(txt, 'Gusts')
    mwind = seri_pasangan_pattern(txt, 'Mean wind speed 10min') or seri_pasangan_pattern(txt, 'Mean wind')
    press = seri_pasangan_pattern(txt, 'Pressure (MSL)') or seri_pasangan_pattern(txt, 'Pressure')
    hum = seri_pasangan_pattern(txt, 'Humidity')

    # Probability of precipitation series
    prob_01 = seri_pasangan_pattern(txt, 'Probability of precipitation &gt;0.1mm')
    prob_1 = seri_pasangan_pattern(txt, 'Probability of precipitation &gt;1mm')
    prob_10 = seri_pasangan_pattern(txt, 'Probability of precipitation &gt;10mm')

    # Columns series
    rain = seri_kolom(txt, 'Rain') or seri_pasangan(txt, 'Rain') or seri_kolom(txt, 'Precipitation sum')
    sun = seri_kolom(txt, 'Sunshine')

    # Cloud layers
    clouds = seri_awan(txt)

    # Arrays
    symbols = seri_array(txt, 'hccompact_data_symbols')
    direction = seri_array(txt, 'hccompact_data_direction')

    if not temp:
        raise RuntimeError(f"Temperature series not found for model: {model_name}")

    ts = sorted(temp.keys())
    rows = []
    for i, t in enumerate(ts):
        dt = datetime.fromtimestamp(t / 1000, WITA)
        
        # Determine wind direction
        deg = None
        if i < len(direction):
            try:
                deg = float(direction[i])
            except (ValueError, TypeError):
                deg = None
                
        sym_raw = symbols[i] if i < len(symbols) else ''
        mw = mwind.get(t)
        gs = gust.get(t)
        c_dict = clouds.get(t, {0: 0.0, 1: 0.0, 2: 0.0})
        
        rows.append({
            'Datetime': dt.strftime('%Y-%m-%d %H:%M:%S'),
            'Hours': dt.hour,
            'Temperature': temp.get(t),
            'Dewpoint': dew.get(t),
            'Humidity': hum.get(t),
            'Pressure': press.get(t),
            'Rain': rain.get(t, 0.0),
            'Prob_Precip_0.1': prob_01.get(t, 0.0),
            'Prob_Precip_1.0': prob_1.get(t, 0.0),
            'Prob_Precip_10.0': prob_10.get(t, 0.0),
            # Wind speed is converted from km/h to knots
            'Wind': (round(mw / KPH_PER_KNOT, 1) if mw is not None else 0.0),
            'Gust': (round(gs / KPH_PER_KNOT, 1) if gs is not None else 0.0),
            'Wind Dir.': deg,
            'Sunshine': sun.get(t, 0.0),
            'Low_Clouds': c_dict.get(0, 0.0),
            'Mid_Clouds': c_dict.get(1, 0.0),
            'High_Clouds': c_dict.get(2, 0.0),
            'Scraped_At': datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S'),
            'Run_Init_UTC': run_init_utc.strftime('%Y-%m-%d %H:%M:%S') if run_init_utc else None,
            'Condition': SIMBOL.get(sym_raw, sym_raw),
            'Model': model_name,
            'Location': STATION_NAME
        })
    return rows

# ================== EXCEL & PARQUET SAVER ==================

def write_excel_sheets(all_models_data: dict, excel_path: str):
    wb = Workbook()
    wb.remove(wb.active) # Remove default sheet
    
    font_header = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    fill_header = PatternFill("solid", fgColor="1F3864")
    align_center = Alignment(horizontal="center", vertical="center")
    border_thin = Border(*([Side(style="thin", color="D0D0D0")] * 4))
    fill_rain = PatternFill("solid", fgColor="DDEBF7")
    
    headers = [
        "Datetime", "Hours", "Temperature (°C)", "Dewpoint (°C)", 
        "Humidity (%)", "Pressure (hPa)", "Rain (mm)", 
        "Prob. Rain >0.1mm (%)", "Prob. Rain >1.0mm (%)", "Prob. Rain >10.0mm (%)", 
        "Wind (kt)", "Gust (kt)", "Wind Dir. (°)", "Sunshine (min)", 
        "Low Clouds (%)", "Mid Clouds (%)", "High Clouds (%)", "Condition"
    ]
    keys = [
        "Datetime", "Hours", "Temperature", "Dewpoint", 
        "Humidity", "Pressure", "Rain", 
        "Prob_Precip_0.1", "Prob_Precip_1.0", "Prob_Precip_10.0", 
        "Wind", "Gust", "Wind Dir.", "Sunshine", 
        "Low_Clouds", "Mid_Clouds", "High_Clouds", "Condition"
    ]

    for model, rows in all_models_data.items():
        ws = wb.create_sheet(title=model)
        ws.cell(row=1, column=1, value=f"Location: {STATION_NAME} | Model: {model}").font = Font(bold=True)
        
        # Write headers
        ws.append([]) # Empty row 2
        ws.append([]) # Empty row 3
        ws.append(headers) # Row 4
        
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=4, column=col_idx)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = border_thin
            
        # Write rows
        for r in rows:
            ws.append([r.get(k) for k in keys])
            
        # Formatting rows
        for r_idx in range(5, len(rows) + 5):
            for c_idx in range(1, len(keys) + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.border = border_thin
                cell.alignment = align_center
                cell.font = Font(name="Arial", size=10)
            
            # Format numbers
            for col in (3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 14, 15, 16, 17):
                cell = ws.cell(row=r_idx, column=col)
                if cell.value is not None:
                    cell.number_format = "0.0"
            ws.cell(row=r_idx, column=13).number_format = "0" # Direction
            
            # Highlight rainy cells (if rain > 0 or precipitation probability >= 50%)
            rain_val = ws.cell(row=r_idx, column=7).value
            prob_val = ws.cell(row=r_idx, column=8).value
            is_rainy = (rain_val and float(rain_val) > 0) or (prob_val and float(prob_val) >= 50.0)
            if is_rainy:
                for c_idx in range(1, len(keys) + 1):
                    ws.cell(row=r_idx, column=c_idx).fill = fill_rain
                    
        # Auto-adjust column widths
        widths = [20, 8, 15, 15, 12, 14, 12, 20, 20, 20, 11, 11, 13, 14, 14, 14, 14, 20]
        from openpyxl.utils import get_column_letter
        for idx, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = w
            
    wb.save(excel_path)

# ================== MAIN EXECUTION ==================

def save_outputs(all_models_data: dict, stacked_rows: list, folder: str, stamp: str):
    """Save Excel, Parquet sidecar, and JSON payloads (optional step)."""
    # 1. Save Excel
    excel_path = os.path.join(folder, f"Prakiraan_Meteologix_Multi_{stamp}.xlsx")
    try:
        write_excel_sheets(all_models_data, excel_path)
        log.info("Excel Saved: %s", excel_path)
    except Exception as e:
        log.error("Failed to write Excel: %s", e)
        
    # 2. Save flat Parquet sidecar
    parquet_path = excel_path.replace(".xlsx", ".parquet")
    try:
        df = pd.DataFrame(stacked_rows)
        df_sys = df.rename(columns={'Dewpoint': 'Dew Point'}).copy()
        
        for col in [
            'Hours', 'Temperature', 'Dew Point', 'Humidity', 'Pressure', 
            'Rain', 'Prob_Precip_0.1', 'Prob_Precip_1.0', 'Prob_Precip_10.0', 
            'Wind', 'Gust', 'Wind Dir.', 'Sunshine', 'Low_Clouds', 'Mid_Clouds', 'High_Clouds'
        ]:
            if col in df_sys.columns:
                df_sys[col] = pd.to_numeric(df_sys[col], errors='coerce')
                
        df_sys.to_parquet(parquet_path, index=False, engine='pyarrow')
        log.info("Parquet Sidecar Saved: %s", parquet_path)
    except Exception as e:
        log.error("Failed to write Parquet sidecar: %s", e)
        
    # 3. Save JSON payloads
    payload = {
        "diperbarui": datetime.now(WITA).strftime("%Y-%m-%d %H:%M:%S WITA"),
        "sumber": "meteologix.com",
        "stasiun": "WAWP - Sangia Nibandera",
        "models_included": list(all_models_data.keys()),
        "data": stacked_rows
    }
    try:
        json_harian = os.path.join(folder, f"Prakiraan_Meteologix_Multi_{stamp}.json")
        with open(json_harian, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
            
        json_terbaru = os.path.join(ARCHIVE_DIR, "prakiraan_multi_terbaru.json")
        with open(json_terbaru, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
            
        log.info("JSON Payloads Saved: %s (+ prakiraan_multi_terbaru.json)", json_harian)
    except Exception as e:
        log.warning("Failed to write JSON: %s", e)

# ================== MAIN EXECUTION ==================

def main() -> tuple[dict, list]:
    log.info("Starting Meteologix Multi-Model Scraping (7 Models)...")
    stamp = datetime.now(WITA).strftime("%Y%m%d_%H%M")
    date_str = datetime.now(WITA).strftime("%Y%m%d")
    
    folder = os.path.join(ARCHIVE_DIR, date_str)
    os.makedirs(folder, exist_ok=True)
    
    all_models_data = {}
    stacked_rows = []
    
    for model_name, url in MODEL_URLS.items():
        log.info("Processing model: %s", model_name)
        try:
            script_txt, run_info_text, update_info_text = scrape_meteogram_script(url)
            
            # Prioritize HTML div text for init time; fallback to JS variable
            run_init_utc = parse_run_init(run_info_text) if run_info_text else None
            if run_init_utc is None:
                run_init_utc = parse_run_init(script_txt)
                if run_init_utc:
                    log.warning("  Used JS fallback for init time (HTML div missing)")
            
            if run_init_utc:
                log.info("  Init time: %s", run_init_utc.strftime("%Y-%m-%d %H:%MZ"))
            else:
                log.warning("  Could not determine init time for %s", model_name)
            
            model_rows = decode_model_script(script_txt, model_name, run_init_utc)
            all_models_data[model_name] = model_rows
            stacked_rows.extend(model_rows)
            log.info("  ✓ Successfully parsed %d forecast points.", len(model_rows))
            
        except Exception as e:
            log.error("  %s: %s", model_name, e)
            
    if not all_models_data:
        log.error("No model data could be scraped. Exiting.")
        sys.exit(1)
        
    return all_models_data, stacked_rows

if __name__ == "__main__":
    all_models_data, stacked_rows = main()
    # For standalone testing, save outputs
    stamp = datetime.now(WITA).strftime("%Y%m%d_%H%M")
    date_str = datetime.now(WITA).strftime("%Y%m%d")
    folder = os.path.join(ARCHIVE_DIR, date_str)
    save_outputs(all_models_data, stacked_rows, folder, stamp)
