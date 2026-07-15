"""Create a leader-ready, fully auditable June 2026 original-TAF metric PDF."""
from __future__ import annotations

import argparse
import csv
import json
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    KeepTogether,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)


PAGE_SIZE = landscape(A4)
NAVY = colors.HexColor("#18344F")
BLUE = colors.HexColor("#236B8E")
LIGHT_BLUE = colors.HexColor("#EAF3F8")
LIGHT_GRAY = colors.HexColor("#F3F5F7")
GREEN = colors.HexColor("#DDEFE2")
RED = colors.HexColor("#F9E2E0")
AMBER = colors.HexColor("#F9EDD7")


def _read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def _is_true(value: Any) -> bool:
    return str(value).strip().lower() == "true"


def _hour_class(row: dict[str, str]) -> str:
    forecast = _is_true(row["forecast_rain"])
    observed = _is_true(row["observed_rain"])
    if forecast and observed:
        return "HIT"
    if forecast:
        return "FALSE ALARM"
    if observed:
        return "MISS"
    return "CORRECT DRY"


def _read_workbook_score(workbook_path: Path) -> dict[str, Any]:
    formula_book = load_workbook(workbook_path, read_only=False, data_only=False)
    value_book = load_workbook(workbook_path, read_only=False, data_only=True)
    try:
        formula_sheet = formula_book["rekap 1 bulan"]
        value_sheet = value_book["rekap 1 bulan"]
        return {
            "monthly_value": float(value_sheet["L14"].value),
            "monthly_formula": str(formula_sheet["L14"].value),
            "issuance_formula": str(formula_sheet["L17"].value),
            "worksheet_endapan_formula": str(formula_book["1"]["W64"].value),
            "hourly_endapan_formula": str(formula_book["1"]["W7"].value),
            "first_day_value": float(value_sheet["L17"].value),
        }
    finally:
        formula_book.close()
        value_book.close()


def _styles() -> dict[str, ParagraphStyle]:
    base = getSampleStyleSheet()
    return {
        "title": ParagraphStyle("title", parent=base["Title"], fontName="Helvetica-Bold", fontSize=22, leading=26, textColor=NAVY, alignment=TA_LEFT, spaceAfter=8),
        "subtitle": ParagraphStyle("subtitle", parent=base["Normal"], fontName="Helvetica", fontSize=10, leading=14, textColor=colors.HexColor("#4C6171"), spaceAfter=12),
        "h1": ParagraphStyle("h1", parent=base["Heading1"], fontName="Helvetica-Bold", fontSize=15, leading=19, textColor=NAVY, spaceBefore=10, spaceAfter=6),
        "h2": ParagraphStyle("h2", parent=base["Heading2"], fontName="Helvetica-Bold", fontSize=11, leading=14, textColor=NAVY, spaceBefore=8, spaceAfter=4),
        "body": ParagraphStyle("body", parent=base["BodyText"], fontName="Helvetica", fontSize=8.5, leading=11.5, textColor=colors.HexColor("#24313D"), spaceAfter=5),
        "small": ParagraphStyle("small", parent=base["BodyText"], fontName="Helvetica", fontSize=7.2, leading=9, textColor=colors.HexColor("#334553")),
        "table": ParagraphStyle("table", parent=base["BodyText"], fontName="Helvetica", fontSize=6.4, leading=7.5, textColor=colors.HexColor("#1F2F3C")),
        "table_bold": ParagraphStyle("table_bold", parent=base["BodyText"], fontName="Helvetica-Bold", fontSize=6.4, leading=7.5, textColor=colors.white, alignment=TA_CENTER),
    }


def _paragraph(text: Any, style: ParagraphStyle) -> Paragraph:
    return Paragraph(str(text).replace("&", "&amp;"), style)


def _footer(canvas, document) -> None:
    canvas.saveState()
    canvas.setStrokeColor(colors.HexColor("#C7D3DC"))
    canvas.line(15 * mm, 12 * mm, PAGE_SIZE[0] - 15 * mm, 12 * mm)
    canvas.setFont("Helvetica", 7)
    canvas.setFillColor(colors.HexColor("#526978"))
    canvas.drawString(15 * mm, 7.5 * mm, "WAWP June 2026 Original TAF Metric Audit - experimental verification evidence")
    canvas.drawRightString(PAGE_SIZE[0] - 15 * mm, 7.5 * mm, f"Page {document.page}")
    canvas.restoreState()


def _metric_table(scores: dict[str, Any], styles: dict[str, ParagraphStyle]) -> Table:
    rain = scores["events"]["rain_any"]
    data = [
        [_paragraph("Metric", styles["table_bold"]), _paragraph("Value", styles["table_bold"]), _paragraph("Meaning", styles["table_bold"])],
        ["Shared official issuances", str(scores["issuance_count"]), "Only issuances with a complete machine-replay counterpart."],
        ["Scored hourly slots", str(rain["sample_size"]), "116 issuances x 24 validity hours."],
        ["Observed rain hours", str(rain["hourly_hits"] + rain["hourly_misses"]), "Explicit RA/SHRA/TSRA in hourly METAR text."],
        ["Exact-hour hits", str(rain["hourly_hits"]), "Forecast and METAR both reported rain in the same hour."],
        ["Exact-hour misses", str(rain["hourly_misses"]), "METAR reported rain but original TAF did not in that hour."],
        ["Exact-hour false alarms", str(rain["hourly_false_alarms"]), "Original TAF reported rain but METAR did not in that hour."],
        ["Exact-hour correct dry", str(rain["hourly_correct_negatives"]), "Both original TAF and METAR were dry."],
        ["All-hour accuracy", f"{rain['all_hour_accuracy'] * 100:.2f}%", "Exact wet/dry agreement across every scored hour."],
        ["Event POD", f"{rain['POD'] * 100:.2f}%", "One-to-one rain-event detection within +/-2 hours."],
        ["Event FAR", f"{rain['FAR'] * 100:.2f}%", "Share of forecast rain events not matched to observed rain."],
        ["Event CSI", f"{rain['CSI'] * 100:.2f}%", "Threat score balancing hits, misses, and false alarms."],
        ["Balanced accuracy", f"{rain['balanced_accuracy'] * 100:.2f}%", "Average of exact-hour rain detection and dry specificity."],
    ]
    table = Table(data, colWidths=[52 * mm, 33 * mm, 165 * mm], repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), NAVY),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("BACKGROUND", (0, 1), (-1, -1), colors.white),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT_GRAY]),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5DD")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTNAME", (0, 1), (1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 1), (-1, -1), 7.3),
        ("LEADING", (0, 1), (-1, -1), 9),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    return table


def _appendix_table(issuance: str, rows: list[dict[str, str]], taf: str, styles: dict[str, ParagraphStyle]) -> KeepTogether:
    header = [
        _paragraph(f"Issuance {issuance}", styles["h2"]),
        _paragraph(f"TAF: {taf}", styles["small"]),
    ]
    data: list[list[Any]] = [[
        _paragraph("UTC", styles["table_bold"]),
        _paragraph("Forecast wx", styles["table_bold"]),
        _paragraph("METAR", styles["table_bold"]),
        _paragraph("F rain", styles["table_bold"]),
        _paragraph("O rain", styles["table_bold"]),
        _paragraph("Exact class", styles["table_bold"]),
        _paragraph("F TS", styles["table_bold"]),
        _paragraph("O TS", styles["table_bold"]),
    ]]
    for row in sorted(rows, key=lambda value: value["valid_time_utc"]):
        data.append([
            row["valid_time_utc"].replace("T", " ").replace("Z", "")[:16],
            row["forecast_weather"] or "NSW",
            _paragraph(row["metar_text"], styles["table"]),
            "Y" if _is_true(row["forecast_rain"]) else "N",
            "Y" if _is_true(row["observed_rain"]) else "N",
            _hour_class(row),
            "Y" if _is_true(row["forecast_thunderstorm"]) else "N",
            "Y" if _is_true(row["observed_thunderstorm"]) else "N",
        ])
    table = Table(data, colWidths=[28 * mm, 25 * mm, 147 * mm, 14 * mm, 14 * mm, 27 * mm, 12 * mm, 12 * mm], repeatRows=1)
    style_commands = [
        ("BACKGROUND", (0, 0), (-1, 0), NAVY),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.2, colors.HexColor("#CED7DE")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 6.2),
        ("LEADING", (0, 1), (-1, -1), 7.2),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]
    for index, row in enumerate(rows, start=1):
        category = _hour_class(row)
        color = GREEN if category in {"HIT", "CORRECT DRY"} else (AMBER if category == "MISS" else RED)
        style_commands.append(("BACKGROUND", (5, index), (5, index), color))
    table.setStyle(TableStyle(style_commands))
    return KeepTogether(header + [table, Spacer(1, 4 * mm)])


def build_pdf(workbook_path: Path, holdout_dir: Path, output_path: Path) -> None:
    styles = _styles()
    summary = json.loads((holdout_dir / "june_holdout_summary.json").read_text(encoding="utf-8"))
    hourly = [row for row in _read_csv(holdout_dir / "june_holdout_hourly.csv") if row["configuration"] == "human_original"]
    taf_rows = {row["issuance_utc"]: row["taf"] for row in _read_csv(holdout_dir / "june_human_tafs.csv")}
    workbook = _read_workbook_score(workbook_path)
    scores = summary["scores"]["human_original"]
    classes = Counter(_hour_class(row) for row in hourly)
    by_issuance: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in hourly:
        by_issuance[row["issuance_utc"]].append(row)

    document = SimpleDocTemplate(
        str(output_path), pagesize=PAGE_SIZE,
        leftMargin=15 * mm, rightMargin=15 * mm, topMargin=14 * mm, bottomMargin=17 * mm,
        title="WAWP June 2026 Original TAF Metric Audit",
        author="WAWP Forecast Engine",
    )
    story: list[Any] = []
    story.extend([
        Paragraph("WAWP June 2026 Original TAF Metric Audit", styles["title"]),
        Paragraph("Detailed all-hour and event-window verification of the original human TAFs against the June METAR archive", styles["subtitle"]),
        Paragraph("Purpose", styles["h1"]),
        Paragraph("This report explains exactly how the workbook Endapan score is calculated, why it can be high during a mostly dry month, and how that differs from rain-event performance metrics. The appendix contains every scored hourly comparison for the original June TAFs.", styles["body"]),
        Paragraph("Executive Finding", styles["h1"]),
        Paragraph(
            f"The workbook reports {workbook['monthly_value']:.2f}% Endapan accuracy. The independent shared-window audit reports {scores['events']['rain_any']['all_hour_accuracy'] * 100:.2f}% exact-hour rain accuracy across {scores['issuance_count']} official issuances and {scores['events']['rain_any']['sample_size']} validity hours. Both are high because dry/dry matches dominate. Rain-event POD is {scores['events']['rain_any']['POD'] * 100:.2f}%, showing that all-hour accuracy and rain-event detection answer different questions.",
            styles["body"],
        ),
        Paragraph("Data Scope and Audit Boundary", styles["h1"]),
    ])
    scope = [
        ["Source workbook", workbook_path.name],
        ["TAF source sheet", "bank data taf"],
        ["METAR source sheet", "bank data metar"],
        ["Human TAFs extracted", str(summary["human_taf_count"])],
        ["Official-slot human TAFs", str(summary["official_human_taf_count"])],
        ["Shared official issuances audited", str(summary["common_issuance_count"])],
        ["Hourly METAR observations", str(summary["hourly_metar_count"])],
        ["Excluded nonstandard human issuance", ", ".join(summary["nonstandard_human_issuances"]) or "None"],
        ["Excluded archive-boundary official issuances", ", ".join(summary["skipped_replay_issuances"]) or "None"],
        ["Rain observation definition", "Explicit RA, SHRA, or TSRA wording in hourly METAR text"],
    ]
    scope_table = Table(scope, colWidths=[68 * mm, 182 * mm])
    scope_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), LIGHT_BLUE),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5DD")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("LEADING", (0, 0), (-1, -1), 10),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.extend([scope_table, Spacer(1, 4 * mm), Paragraph("Workbook Metric Calculation", styles["h1"]), Paragraph("The workbook Endapan number is a layered average of equality checks, not an event-detection score.", styles["body"])])
    formula_data = [
        ["Level", "Workbook evidence", "Interpretation"],
        ["Monthly rekap", workbook["monthly_formula"], "L14 averages 120 individual TAF Endapan percentages."],
        ["Per-TAF link", workbook["issuance_formula"], "Each row in rekap links to a daily-sheet Endapan result."],
        ["Daily Endapan score", workbook["worksheet_endapan_formula"], "The daily sheet sums Endapan matches and divides by its populated hourly cells."],
        ["Hourly Endapan check", workbook["hourly_endapan_formula"], "An hour scores 1 only when forecast Endapan equals observed Endapan; otherwise 0."],
    ]
    formula_table = Table([[_paragraph(value, styles["table_bold"]) for value in formula_data[0]]] + [[_paragraph(value, styles["small"]) for value in row] for row in formula_data[1:]], colWidths=[45 * mm, 102 * mm, 103 * mm], repeatRows=1)
    formula_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), NAVY),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT_GRAY]),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5DD")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.extend([formula_table, Spacer(1, 4 * mm), Paragraph("Independent Metric Calculation", styles["h1"]), Paragraph("The independent audit applies two complementary views to the same original TAFs: exact-hour wet/dry agreement, and one-to-one rain-event matching within +/-2 hours. The first is comparable in spirit to the workbook Endapan score; the second tests whether observed rain events were actually anticipated.", styles["body"]), _metric_table(scores, styles), PageBreak()])
    story.extend([
        Paragraph("Why 98.68% Accuracy and 33.33% POD Can Both Be True", styles["h1"]),
        Paragraph("Rain is uncommon in the verification sample. A forecast can receive a high all-hour accuracy score by correctly describing the many dry hours, even when it misses a substantial share of the fewer wet hours. POD excludes correct dry hours from its denominator; accuracy includes them.", styles["body"]),
    ])
    reconciliation = [
        ["Measure", "Numerator / denominator", "Result", "Use"],
        ["Workbook Endapan", "Average of linked per-TAF equality percentages", f"{workbook['monthly_value']:.2f}%", "Operational workbook reference"],
        ["Independent all-hour accuracy", f"({classes['HIT']} exact wet hits + {classes['CORRECT DRY']} correct dry) / {len(hourly)} hours", f"{scores['events']['rain_any']['all_hour_accuracy'] * 100:.2f}%", "Exact shared-window wet/dry agreement"],
        ["Rain-event POD", f"{scores['events']['rain_any']['hits']} event-window hits / {scores['events']['rain_any']['observed_events']} observed rain events", f"{scores['events']['rain_any']['POD'] * 100:.2f}%", "Rain-event detection within +/-2 hours"],
        ["Rain-event CSI", "Hits / (hits + misses + false alarms)", f"{scores['events']['rain_any']['CSI'] * 100:.2f}%", "Threat score that penalizes both misses and false alarms"],
    ]
    recon_table = Table([[_paragraph(value, styles["table_bold"]) for value in reconciliation[0]]] + [[_paragraph(value, styles["small"]) for value in row] for row in reconciliation[1:]], colWidths=[47 * mm, 110 * mm, 28 * mm, 65 * mm], repeatRows=1)
    recon_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), NAVY),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT_GRAY]),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5DD")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.extend([recon_table, Spacer(1, 4 * mm), Paragraph("Interpretation", styles["h1"]), Paragraph("The small difference between the workbook 98.68% and independent 97.41% is expected. The workbook averages 120 source entries using its daily template and special TEMPO aggregation, while the independent audit uses only the 116 official issuances shared with a complete archived machine forecast and applies a direct hourly METAR-text weather definition. The scores are similar in direction and both reflect dry-hour dominance; neither should be used alone to assess rare rain-event skill.", styles["body"]), Paragraph("Hourly Classification Legend", styles["h1"]), Paragraph("CORRECT DRY: both TAF and METAR were dry. HIT: both reported rain in the exact hour. MISS: METAR reported rain but the TAF did not. FALSE ALARM: TAF reported rain but METAR did not. The appendix retains all hours, not only errors.", styles["body"]), PageBreak(), Paragraph("Appendix A - Complete Hourly Original TAF Audit", styles["h1"]), Paragraph("Each table shows the original TAF, its active structured weather state at every validity hour, the paired hourly METAR, and exact rain/thunderstorm classification. The detail is limited to the 116 official TAF issuance windows that have complete shared archival coverage.", styles["body"]), Spacer(1, 3 * mm)])
    for issuance in sorted(by_issuance):
        story.append(_appendix_table(issuance, by_issuance[issuance], taf_rows.get(issuance, "Source TAF unavailable"), styles))
        story.append(PageBreak())
    document.build(story, onFirstPage=_footer, onLaterPages=_footer)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, default=Path(r"C:\Users\MY ASUS\Downloads\Verifikasi TAF_FORM_Juni_2026 (2).xlsx"))
    parser.add_argument("--holdout-dir", type=Path, default=Path(r"D:\UJI_PERFORMA_MODEL\VERIFICATION_REPORTS\taf_june_holdout_2026"))
    parser.add_argument("--output", type=Path, default=Path(r"D:\UJI_PERFORMA_MODEL\meteologix-wawp-main\output\pdf\WAWP_June_2026_Original_TAF_Metric_Audit.pdf"))
    args = parser.parse_args()
    args.output.parent.mkdir(parents=True, exist_ok=True)
    build_pdf(args.workbook, args.holdout_dir, args.output)
    print(args.output)


if __name__ == "__main__":
    main()
