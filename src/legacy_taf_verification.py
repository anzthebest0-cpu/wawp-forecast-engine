"""Independent, auditable verification for the historical workbook TAF inputs.

The original verification workbooks are useful source evidence, but their
formula cache contains Google Sheets-only functions and a few stale values.
This module reads the active ``TAF!C`` source column and scores it directly
against the standalone canonical METAR extracts.  It has two views:

``legacy_compatibility``
    Reproduces the workbook's category rules and unusual TEMPO aggregation as
    closely as possible, including the strict ``> 4 mm`` rain-occurrence gate.

``quality_gated``
    Uses the same category rules, but excludes non-hourly observations,
    invalid reconstructed timestamps, and rain records whose source timestamp
    does not match the observed METAR time.  This is the defensible view for
    comparing the historical pre-configured TAFs.

The module deliberately does not modify the original Excel workbooks.
"""
from __future__ import annotations

import argparse
import csv
import json
import re
from dataclasses import asdict, dataclass, replace
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


UTC = timezone.utc
METRICS = (
    "wind_direction",
    "wind_speed",
    "wind_gust",
    "visibility",
    "rain_occurrence",
    "cloud_amount",
    "cloud_base",
)
INVALID_OBSERVATION_STATUSES = {"invalid_calendar", "invalid_day_for_verified_period"}

WORKBOOKS = (
    ("2026-01", "01_Verifikasi_Taf.xlsx"),
    ("2026-02", "02_Verifikasi_Taf (1).xlsx"),
    ("2026-03", "03_Verifikasi_Taf (1).xlsx"),
    ("2026-04", "04_Verifikasi_Taf.xlsx"),
    ("2026-05", "05_Verifikasi_Taf.xlsx"),
)

WIND_RE = re.compile(r"\b(?P<direction>\d{3}|VRB)(?P<speed>\d{2})(?:G(?P<gust>\d{2}))?KT\b")
VIS_RE = re.compile(r"\b(\d{4})\b")
CLOUD_RE = re.compile(r"\b(FEW|SCT|BKN|OVC)(\d{3})(?:CB|TCU)?\b")
WX_RE = re.compile(
    r"\b(?:-?DZ|-?RA|\+RA|-?SN|-?SG|-?IC|-?PL|-?GR|-?GS|-?UP|-?BR|-?FG|-?FU|-?VA|-?DU|-?SA|-?HZ|-?PY|-?PO|-?SQ|-?FC|-?SS|-?DS|MIFG|BCFG|PRFG|DRSN|BLSN|-?SHRA|-?TSRA|-?FZRA|-?SHSN|-?TSSN|-?FZSN|-?SHGR|-?TSGR|-?FZGR|-?SHGS|-?TSGS|-?FZGS)\b"
)
GROUP_RE = re.compile(r"\b(?P<kind>BECMG|TEMPO|PROB\d{2}\s+TEMPO)\s+(?P<window>\d{4}/\d{4})\b")
HEADER_RE = re.compile(r"\bTAF\s+WAWP\s+(?P<issue>\d{6})Z\s+(?P<valid>\d{4}/\d{4})\b")


@dataclass(frozen=True)
class WeatherState:
    wind_direction: int | None
    wind_is_variable: bool
    wind_speed: int | None
    wind_gust: int | None
    visibility_m: int | None
    weather: str | None
    cloud_amount: str | None
    cloud_base_ft: int | None


@dataclass(frozen=True)
class ChangeGroup:
    kind: str
    start: datetime
    end: datetime
    remainder: str
    has_wind: bool


@dataclass(frozen=True)
class ParsedTAF:
    text: str
    issue_time: datetime
    valid_start: datetime
    valid_end: datetime
    base: WeatherState
    groups: tuple[ChangeGroup, ...]


def _month_shift(year: int, month: int, delta: int) -> tuple[int, int]:
    month += delta
    while month < 1:
        year -= 1
        month += 12
    while month > 12:
        year += 1
        month -= 12
    return year, month


def _day_candidates(year: int, month: int, day: int, hour: int) -> list[datetime]:
    """Return valid UTC candidates in the surrounding three calendar months."""
    candidates: list[datetime] = []
    for delta in (0, 1, -1):
        candidate_year, candidate_month = _month_shift(year, month, delta)
        try:
            candidates.append(datetime(candidate_year, candidate_month, day, hour, tzinfo=UTC))
        except ValueError:
            continue
    if not candidates:
        raise ValueError(f"Cannot resolve TAF day {day:02d} around {year:04d}-{month:02d}")
    return candidates


def _resolve_group_time(
    year: int,
    month: int,
    start_day: int,
    start_hour: int,
    end_day: int,
    end_hour: int,
    anchor: datetime | None = None,
) -> tuple[datetime, datetime]:
    reference = anchor or datetime(year, month, start_day, start_hour, tzinfo=UTC)
    start = min(
        _day_candidates(year, month, start_day, start_hour),
        key=lambda value: abs((value - reference).total_seconds()),
    )
    end_candidates = [
        value for value in _day_candidates(year, month, end_day, end_hour) if value > start
    ]
    if not end_candidates:
        raise ValueError(
            f"TAF end {end_day:02d}{end_hour:02d} is not after start "
            f"{start_day:02d}{start_hour:02d} around {year:04d}-{month:02d}"
        )
    end = min(end_candidates)
    return start, end


def _parse_state(segment: str) -> WeatherState:
    wind = WIND_RE.search(segment)
    wind_direction = None
    wind_is_variable = False
    wind_speed = None
    wind_gust = None
    remainder = segment
    if wind:
        direction = wind.group("direction")
        wind_is_variable = direction == "VRB"
        wind_direction = None if wind_is_variable else int(direction)
        wind_speed = int(wind.group("speed"))
        wind_gust = int(wind.group("gust") or 0)
        remainder = segment[wind.end() :]

    visibility_m = 9999 if "CAVOK" in remainder else None
    if visibility_m is None:
        vis = VIS_RE.search(remainder)
        visibility_m = int(vis.group(1)) if vis else None

    weather_match = WX_RE.search(remainder)
    weather = weather_match.group(0) if weather_match else None
    cloud = CLOUD_RE.search(remainder)
    cloud_amount = cloud.group(1) if cloud else ("SKC" if "CAVOK" in remainder else None)
    cloud_base_ft = int(cloud.group(2)) * 100 if cloud else (9999 if "CAVOK" in remainder else None)
    return WeatherState(
        wind_direction=wind_direction,
        wind_is_variable=wind_is_variable,
        wind_speed=wind_speed,
        wind_gust=wind_gust,
        visibility_m=visibility_m,
        weather=weather,
        cloud_amount=cloud_amount,
        cloud_base_ft=cloud_base_ft,
    )


def _overlay_state(base: WeatherState, override: WeatherState) -> WeatherState:
    """Apply only the elements explicitly supplied by a TAF change group."""
    has_wind = override.wind_speed is not None
    return WeatherState(
        wind_direction=override.wind_direction if has_wind else base.wind_direction,
        wind_is_variable=override.wind_is_variable if has_wind else base.wind_is_variable,
        wind_speed=override.wind_speed if has_wind else base.wind_speed,
        wind_gust=override.wind_gust if has_wind else base.wind_gust,
        visibility_m=override.visibility_m if override.visibility_m is not None else base.visibility_m,
        weather=override.weather if override.weather is not None else base.weather,
        cloud_amount=override.cloud_amount if override.cloud_amount is not None else base.cloud_amount,
        cloud_base_ft=override.cloud_base_ft if override.cloud_base_ft is not None else base.cloud_base_ft,
    )


def parse_taf(text: str, period: str) -> ParsedTAF:
    """Parse a source TAF using the target verification period for dates."""
    compact = " ".join(str(text).replace("\r", " ").replace("\n", " ").split())
    header = HEADER_RE.search(compact)
    if not header:
        raise ValueError(f"Malformed WAWP TAF: {text!r}")
    year, month = (int(value) for value in period.split("-"))
    valid_start_day, valid_start_hour = int(header.group("valid")[:2]), int(header.group("valid")[2:4])
    valid_end_day, valid_end_hour = int(header.group("valid")[5:7]), int(header.group("valid")[7:9])
    valid_start, valid_end = _resolve_group_time(
        year,
        month,
        valid_start_day,
        valid_start_hour,
        valid_end_day,
        valid_end_hour,
        anchor=datetime(year, month, valid_start_day, valid_start_hour, tzinfo=UTC),
    )

    issue_day, issue_hour, issue_minute = (
        int(header.group("issue")[:2]),
        int(header.group("issue")[2:4]),
        int(header.group("issue")[4:6]),
    )
    issue_candidates = []
    for delta in (-1, 0, 1):
        issue_year, issue_month = _month_shift(year, month, delta)
        try:
            issue_candidates.append(
                datetime(issue_year, issue_month, issue_day, issue_hour, issue_minute, tzinfo=UTC)
            )
        except ValueError:
            continue
    issue_time = min(issue_candidates, key=lambda value: abs((valid_start - value).total_seconds()))

    first_group = GROUP_RE.search(compact, header.end())
    base_end = first_group.start() if first_group else len(compact)
    parsed_base = _parse_state(compact[header.end() : base_end])
    base = WeatherState(
        wind_direction=parsed_base.wind_direction,
        wind_is_variable=parsed_base.wind_is_variable,
        wind_speed=parsed_base.wind_speed,
        wind_gust=parsed_base.wind_gust,
        visibility_m=parsed_base.visibility_m,
        weather=parsed_base.weather or "0",
        cloud_amount=parsed_base.cloud_amount,
        cloud_base_ft=parsed_base.cloud_base_ft,
    )
    matches = list(GROUP_RE.finditer(compact, header.end()))
    groups: list[ChangeGroup] = []
    for index, match in enumerate(matches):
        next_start = matches[index + 1].start() if index + 1 < len(matches) else len(compact)
        start_day, start_hour = int(match.group("window")[:2]), int(match.group("window")[2:4])
        end_day, end_hour = int(match.group("window")[5:7]), int(match.group("window")[7:9])
        start, end = _resolve_group_time(
            year,
            month,
            start_day,
            start_hour,
            end_day,
            end_hour,
            anchor=valid_start,
        )
        remainder = compact[match.end() : next_start]
        groups.append(
            ChangeGroup(
                kind="TEMPO" if "TEMPO" in match.group("kind") else "BECMG",
                start=start,
                end=end,
                remainder=remainder,
                has_wind=bool(WIND_RE.search(remainder)),
            )
        )
    return ParsedTAF(compact, issue_time, valid_start, valid_end, base, tuple(groups))


def repair_validity_to_24h(taf: ParsedTAF) -> tuple[ParsedTAF, bool, int]:
    """Repair impossible historical validity windows under a stated 24-hour rule.

    The repair changes only the evaluation window, never the source TAF text.
    The reported end and original duration remain in every audit record.
    """
    reported_hours = int((taf.valid_end - taf.valid_start).total_seconds() // 3600)
    if 18 <= reported_hours <= 30:
        return taf, False, reported_hours
    return replace(taf, valid_end=taf.valid_start + timedelta(hours=24)), True, reported_hours


def active_state(taf: ParsedTAF, observed_at: datetime) -> tuple[str, WeatherState, bool]:
    """Return the legacy change label, active forecast values, and TEMPO wind flag."""
    # The workbook checks TEMPO first.  Keep that precedence for comparison.
    for group in taf.groups:
        if group.kind == "TEMPO" and group.start <= observed_at < group.end:
            return "TEMPO", _overlay_state(taf.base, _parse_state(group.remainder)), group.has_wind
    # A BECMG value becomes active only after its stated completion time.  The
    # original workbook labels that persistent state BECMG rather than GENERAL.
    completed = [group for group in taf.groups if group.kind == "BECMG" and observed_at >= group.end]
    if completed:
        group = completed[-1]
        return "BECMG", _overlay_state(taf.base, _parse_state(group.remainder)), False
    return "GENERAL", taf.base, False


def active_state_structured(taf: ParsedTAF, observed_at: datetime) -> tuple[str, WeatherState, bool]:
    """Evaluate multi-group TAF state without the legacy workbook limitations.

    Completed BECMG groups persist in sequence and an active TEMPO overlays the
    currently prevailing state.  This is used for the replay-configuration
    experiment; ``active_state`` remains available for legacy-workbook parity.
    """
    prevailing = taf.base
    has_completed_becmg = False
    for group in taf.groups:
        if group.kind == "BECMG" and observed_at >= group.end:
            prevailing = _overlay_state(prevailing, _parse_state(group.remainder))
            has_completed_becmg = True
    for group in taf.groups:
        if group.kind == "TEMPO" and group.start <= observed_at < group.end:
            return "TEMPO", _overlay_state(prevailing, _parse_state(group.remainder)), group.has_wind
    return ("BECMG" if has_completed_becmg else "GENERAL"), prevailing, False


def parse_metar(text: str) -> WeatherState:
    compact = " ".join(str(text or "").replace("\r", " ").replace("\n", " ").split())
    parsed = _parse_state(compact)
    return WeatherState(
        wind_direction=parsed.wind_direction,
        wind_is_variable=parsed.wind_is_variable,
        wind_speed=parsed.wind_speed,
        wind_gust=parsed.wind_gust,
        visibility_m=parsed.visibility_m,
        weather=parsed.weather or "0",
        cloud_amount=parsed.cloud_amount,
        cloud_base_ft=parsed.cloud_base_ft,
    )


def _visibility_band(value: int | None, legacy_missing_high: bool) -> str | None:
    if value is None:
        return "high" if legacy_missing_high else None
    if value < 800:
        return "lt_800"
    if value <= 1500:
        return "800_1500"
    if value <= 3000:
        return "1500_3000"
    if value <= 5000:
        return "3000_5000"
    return "high"


def _cloud_class(value: str | None) -> str | None:
    if value in {"SKC", "FEW", "SCT"}:
        return "few_sct"
    if value in {"BKN", "OVC"}:
        return "bkn_ovc"
    return None


def score_hour(
    forecast: WeatherState,
    observed: WeatherState,
    observed_rain: str,
    *,
    legacy_missing_visibility_high: bool,
    observed_metar_text: str = "",
) -> dict[str, int | None]:
    """Score one forecast/observation pair under the workbook category rules."""
    scores: dict[str, int | None] = {metric: None for metric in METRICS}
    if forecast.wind_speed is not None and observed.wind_speed is not None:
        if forecast.wind_is_variable:
            direction_hit = "TS" in observed_metar_text or "CB" in observed_metar_text or observed.wind_speed < 10
        elif observed.wind_is_variable:
            direction_hit = forecast.wind_speed < 10 and observed.wind_speed < 10
        elif forecast.wind_direction is not None and observed.wind_direction is not None:
            difference = abs(forecast.wind_direction - observed.wind_direction)
            direction_hit = difference < 61 or (difference > 60 and observed.wind_speed < 10)
        else:
            direction_hit = False
        scores["wind_direction"] = int(direction_hit)
        scores["wind_speed"] = int(abs(forecast.wind_speed - observed.wind_speed) < 11)
    if forecast.wind_gust is not None and observed.wind_gust is not None:
        scores["wind_gust"] = int(forecast.wind_gust == observed.wind_gust)

    forecast_band = _visibility_band(forecast.visibility_m, False)
    observed_band = _visibility_band(observed.visibility_m, legacy_missing_visibility_high)
    if forecast_band is not None and observed_band is not None:
        scores["visibility"] = int(forecast_band == observed_band)
    if forecast.weather is not None:
        scores["rain_occurrence"] = int(forecast.weather == observed_rain)

    forecast_cloud = _cloud_class(forecast.cloud_amount)
    observed_cloud = _cloud_class(observed.cloud_amount)
    if forecast_cloud is not None and observed_cloud is not None:
        scores["cloud_amount"] = int(forecast_cloud == observed_cloud)
    if forecast.cloud_base_ft is not None and observed.cloud_base_ft is not None and forecast.cloud_base_ft > 0:
        if observed.cloud_base_ft < 1001:
            ceiling_hit = abs(forecast.cloud_base_ft - observed.cloud_base_ft) < 101
        else:
            ceiling_hit = abs(observed.cloud_base_ft - forecast.cloud_base_ft) / forecast.cloud_base_ft < 0.30005
        scores["cloud_base"] = int(ceiling_hit)
    return scores


def legacy_tempo_adjustment(hits: int, hours: int, metric: str, tempo_has_wind: bool) -> float:
    """Return the original workbook's intentionally non-monotonic TEMPO credit."""
    if hours == 0:
        return 0.0
    if hits == 0:
        return 0.3 * hours
    if hits < 0.5001 * hours:
        return float(hours)
    if 0.5 * hours < hits < 0.6 * hours:
        return float(hits)
    if metric in {"wind_direction", "wind_speed"} and not tempo_has_wind:
        return float(hits)
    return 0.6 * hours


def _empty_metric_summary() -> dict[str, Any]:
    return {metric: {"eligible_hours": 0, "raw_hits": 0, "legacy_adjusted_hits": 0.0} for metric in METRICS}


def _aggregate_issuance(rows: Iterable[dict[str, Any]], mode: str) -> dict[str, dict[str, float | int | None]]:
    summary = _empty_metric_summary()
    rows_by_metric: dict[str, list[dict[str, Any]]] = {metric: [] for metric in METRICS}
    for row in rows:
        for metric in METRICS:
            value = row.get(f"{mode}_{metric}")
            if value is not None and value != "":
                rows_by_metric[metric].append(row)
    for metric, metric_rows in rows_by_metric.items():
        eligible = len(metric_rows)
        direct_hits = sum(int(row[f"{mode}_{metric}"]) for row in metric_rows if row["change_group"] != "TEMPO")
        tempo_rows = [row for row in metric_rows if row["change_group"] == "TEMPO"]
        tempo_hits = sum(int(row[f"{mode}_{metric}"]) for row in tempo_rows)
        tempo_has_wind = any(bool(row["tempo_has_wind"]) for row in tempo_rows)
        adjusted = direct_hits + legacy_tempo_adjustment(tempo_hits, len(tempo_rows), metric, tempo_has_wind)
        summary[metric] = {
            "eligible_hours": eligible,
            "raw_hits": direct_hits + tempo_hits,
            "legacy_adjusted_hits": adjusted,
            "legacy_score_percent": (100.0 * adjusted / eligible) if eligible else None,
            "raw_score_percent": (100.0 * (direct_hits + tempo_hits) / eligible) if eligible else None,
        }
    return summary


def _as_optional_int(value: Any) -> int | None:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _is_true(value: Any) -> bool:
    return str(value).strip().lower() == "true"


def _read_metar_rows(path: Path) -> list[dict[str, Any]]:
    with path.open("r", newline="", encoding="utf-8") as handle:
        rows = list(csv.DictReader(handle))
    for row in rows:
        raw_time = row.get("observed_at_utc", "")
        row["observed_at"] = datetime.fromisoformat(raw_time.replace("Z", "+00:00")) if raw_time else None
        row["metar_day"] = _as_optional_int(row.get("metar_day_utc"))
        row["metar_hour"] = _as_optional_int(row.get("metar_hour_utc"))
        row["metar_minute"] = _as_optional_int(row.get("metar_minute_utc"))
        row["rainfall_raw_tenths_mm"] = _as_optional_int(row.get("rainfall_raw_tenths_mm"))
    return rows


def _read_tafs(path: Path, period: str) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        sheet = workbook["TAF"]
        tafs = []
        for row_index, values in enumerate(
            sheet.iter_rows(min_row=2, min_col=3, max_col=3, values_only=True), start=2
        ):
            value = values[0]
            if not isinstance(value, str) or "TAF WAWP" not in value:
                continue
            parsed = parse_taf(value, period)
            tafs.append({"source_taf_row": row_index, "taf": parsed})
        return tafs
    finally:
        workbook.close()


def _read_recap(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook["Rekap 1 bulan"]
        cells = {
            "wind_direction": "D14",
            "wind_speed": "F14",
            "wind_gust": "H14",
            "visibility": "J14",
            "rain_occurrence": "L14",
            "cloud_amount": "N14",
            "cloud_base": "P14",
            "workbook_overall": "Q14",
        }
        return {key: sheet[cell].value for key, cell in cells.items()}
    finally:
        workbook.close()


def _legacy_slot_key(row: dict[str, Any], period: str) -> tuple[int, int] | None:
    if row["metar_day"] is None or row["metar_hour"] is None:
        return None
    return row["metar_day"], row["metar_hour"]


def _observed_rain(row: dict[str, Any]) -> str:
    raw = row.get("rainfall_raw_tenths_mm")
    return "RA" if raw is not None and raw / 10.0 > 4.0 else "0"


def _quality_row_is_eligible(row: dict[str, Any]) -> bool:
    return (
        row.get("observed_at") is not None
        and row.get("observed_at_status") not in INVALID_OBSERVATION_STATUSES
        and row.get("metar_minute") == 0
    )


def _quality_rain_provenance(row: dict[str, Any]) -> tuple[bool, str]:
    """Accept a stale date only when the HUJAN hour still matches the METAR row."""
    if _is_true(row.get("rain_source_timestamp_matches_observed_at")):
        return True, "verified_source_timestamp"
    if _is_true(row.get("metar_time_group_matches_rain_timestamp")):
        return True, "timestamp_normalized_from_row_alignment"
    return False, "not_scored_unaligned_rain_timestamp"


def verify_month(
    period: str,
    workbook_path: Path,
    metar_path: Path,
    *,
    structured_taf_state: bool = False,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    """Recalculate one workbook's active TAF column against canonical METAR rows."""
    metar_rows = _read_metar_rows(metar_path)
    taf_rows = _read_tafs(workbook_path, period)
    period_year, period_month = (int(value) for value in period.split("-"))
    legacy_index = {_legacy_slot_key(row, period): row for row in metar_rows if _legacy_slot_key(row, period)}
    quality_index = {row["observed_at"]: row for row in metar_rows if _quality_row_is_eligible(row)}
    state_evaluator = active_state_structured if structured_taf_state else active_state
    hourly: list[dict[str, Any]] = []
    issuance: list[dict[str, Any]] = []
    repaired_validity_rows: list[int] = []

    for item in taf_rows:
        source_taf: ParsedTAF = item["taf"]
        taf, validity_repaired, reported_validity_hours = repair_validity_to_24h(source_taf)
        if validity_repaired:
            repaired_validity_rows.append(item["source_taf_row"])
        issuance_rows: list[dict[str, Any]] = []
        current = taf.valid_start
        while current < taf.valid_end:
            # Source workbook rows belong only to its declared verification
            # month.  Do not let a final-day TAF reuse day 01 from that same
            # month for its next-month validity hours.
            legacy_row = (
                legacy_index.get((current.day, current.hour))
                if (current.year, current.month) == (period_year, period_month)
                else None
            )
            quality_row = quality_index.get(current)
            source_row = legacy_row or quality_row
            change_group, forecast, tempo_has_wind = state_evaluator(taf, current)
            row: dict[str, Any] = {
                "period": period,
                "source_workbook": workbook_path.name,
                "source_taf_row": item["source_taf_row"],
                "taf_issue_utc": taf.issue_time.isoformat().replace("+00:00", "Z"),
                "taf_valid_start_utc": taf.valid_start.isoformat().replace("+00:00", "Z"),
                "taf_valid_end_utc": taf.valid_end.isoformat().replace("+00:00", "Z"),
                "reported_taf_valid_end_utc": source_taf.valid_end.isoformat().replace("+00:00", "Z"),
                "reported_validity_hours": reported_validity_hours,
                "validity_repaired_to_24h": validity_repaired,
                "valid_time_utc": current.isoformat().replace("+00:00", "Z"),
                "taf_text": taf.text,
                "change_group": change_group,
                "tempo_has_wind": tempo_has_wind,
                "forecast_wind_direction": "VRB" if forecast.wind_is_variable else forecast.wind_direction,
                "forecast_wind_speed": forecast.wind_speed,
                "forecast_wind_gust": forecast.wind_gust,
                "forecast_visibility_m": forecast.visibility_m,
                "forecast_weather": forecast.weather,
                "forecast_cloud_amount": forecast.cloud_amount,
                "forecast_cloud_base_ft": forecast.cloud_base_ft,
                "metar_text": source_row.get("metar_text") if source_row else "",
                "metar_time_group": source_row.get("metar_time_group") if source_row else "",
                "source_observation_status": source_row.get("observed_at_status") if source_row else "missing",
                "rain_source_time_matches_observation": source_row.get("rain_source_timestamp_matches_observed_at") if source_row else "",
            }

            legacy_scores = {metric: None for metric in METRICS}
            quality_scores = {metric: None for metric in METRICS}
            if legacy_row:
                observed = parse_metar(legacy_row.get("metar_text", ""))
                legacy_scores = score_hour(
                    forecast,
                    observed,
                    _observed_rain(legacy_row),
                    legacy_missing_visibility_high=True,
                    observed_metar_text=legacy_row.get("metar_text", ""),
                )
                row.update({f"observed_{key}": value for key, value in asdict(observed).items()})
            if quality_row:
                observed = parse_metar(quality_row.get("metar_text", ""))
                quality_scores = score_hour(
                    forecast,
                    observed,
                    _observed_rain(quality_row),
                    legacy_missing_visibility_high=False,
                    observed_metar_text=quality_row.get("metar_text", ""),
                )
                rain_eligible, rain_provenance = _quality_rain_provenance(quality_row)
                row["quality_rain_provenance"] = rain_provenance
                if not rain_eligible:
                    quality_scores["rain_occurrence"] = None
                row.update({f"quality_observed_{key}": value for key, value in asdict(observed).items()})
            row.update({f"legacy_compatibility_{metric}": legacy_scores[metric] for metric in METRICS})
            row.update({f"quality_gated_{metric}": quality_scores[metric] for metric in METRICS})
            hourly.append(row)
            issuance_rows.append(row)
            current += timedelta(hours=1)

        legacy_summary = _aggregate_issuance(issuance_rows, "legacy_compatibility")
        quality_summary = _aggregate_issuance(issuance_rows, "quality_gated")
        issue_record: dict[str, Any] = {
            "period": period,
            "source_workbook": workbook_path.name,
            "source_taf_row": item["source_taf_row"],
            "taf_issue_utc": taf.issue_time.isoformat().replace("+00:00", "Z"),
            "taf_valid_start_utc": taf.valid_start.isoformat().replace("+00:00", "Z"),
            "taf_valid_end_utc": taf.valid_end.isoformat().replace("+00:00", "Z"),
            "reported_taf_valid_end_utc": source_taf.valid_end.isoformat().replace("+00:00", "Z"),
            "reported_validity_hours": reported_validity_hours,
            "validity_repaired_to_24h": validity_repaired,
            "taf_text": taf.text,
            "verification_status": "scored_with_repaired_validity" if validity_repaired else "scored",
            "validity_hours": int((taf.valid_end - taf.valid_start).total_seconds() // 3600),
        }
        for mode, summary in (("legacy_compatibility", legacy_summary), ("quality_gated", quality_summary)):
            for metric in METRICS:
                for field, value in summary[metric].items():
                    issue_record[f"{mode}_{metric}_{field}"] = value
        issuance.append(issue_record)

    recap = _read_recap(workbook_path)
    report = {
        "period": period,
        "source_workbook": workbook_path.name,
        "source_taf_count": len(taf_rows),
        "scored_taf_count": len(taf_rows),
        "repaired_validity_rows": repaired_validity_rows,
        "hourly_rows": len(hourly),
        "source_recap_cached_values": recap,
    }
    return hourly, issuance, report


def _monthly_aggregate(issuance_rows: list[dict[str, Any]], mode: str) -> dict[str, dict[str, float | int | None]]:
    summary: dict[str, dict[str, float | int | None]] = {}
    for metric in METRICS:
        eligible = sum(int(row.get(f"{mode}_{metric}_eligible_hours") or 0) for row in issuance_rows)
        raw_hits = sum(int(row.get(f"{mode}_{metric}_raw_hits") or 0) for row in issuance_rows)
        adjusted_hits = sum(float(row.get(f"{mode}_{metric}_legacy_adjusted_hits") or 0) for row in issuance_rows)
        issuance_legacy_scores = [
            float(row[f"{mode}_{metric}_legacy_score_percent"])
            for row in issuance_rows
            if row.get(f"{mode}_{metric}_legacy_score_percent") is not None
        ]
        issuance_raw_scores = [
            float(row[f"{mode}_{metric}_raw_score_percent"])
            for row in issuance_rows
            if row.get(f"{mode}_{metric}_raw_score_percent") is not None
        ]
        summary[metric] = {
            "eligible_hours": eligible,
            "raw_hits": raw_hits,
            "legacy_adjusted_hits": adjusted_hits,
            # The source monthly recap uses AVERAGE over each issuance's
            # summary cell. Preserve that equal-issuance aggregation here.
            "legacy_score_percent": (
                sum(issuance_legacy_scores) / len(issuance_legacy_scores)
                if issuance_legacy_scores else None
            ),
            "raw_score_percent": (
                sum(issuance_raw_scores) / len(issuance_raw_scores)
                if issuance_raw_scores else None
            ),
            "issuance_count": len(issuance_legacy_scores),
        }
    return summary


def _write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    keys = sorted({key for row in rows for key in row}) if rows else []
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=keys)
        writer.writeheader()
        writer.writerows(rows)


def _format_score(value: Any) -> str:
    return "not scored" if value is None else f"{float(value):.2f}%"


def build_report(months: list[dict[str, Any]], output_dir: Path, *, structured_taf_state: bool = False) -> Path:
    lines = [
        "# Historical Pre-configured TAF Recalculation",
        "",
        "## Scope",
        "",
        "This report independently scores the active `TAF!C` forecasts in the January-May 2026 verification workbooks against the canonical standalone METAR extracts. The original workbooks are not modified.",
        "",
        "Two result views are provided:",
        "",
        "- **Legacy compatibility:** reproduces the original category thresholds, the strict `>4 mm` rain gate, row-slot matching, and the original non-monotonic TEMPO adjustment.",
        "- **Quality gated:** keeps the same scoring categories, excludes non-hourly and invalid-calendar observations, and normalizes stale rain dates only when the HUJAN hour matches the paired METAR hour. Use this view for a defensible result.",
        "",
        "## Results",
        "",
        "| Period | Scored/source TAFs | View | Wind dir. | Wind speed | Gust | Visibility | Rain occurrence | Cloud amount | Cloud base |",
        "| --- | ---: | --- | ---: | ---: | ---: | ---: | ---: | ---: | ---: |",
    ]
    if structured_taf_state:
        lines.extend([
            "",
            "The report uses the structured TAF state evaluator: completed BECMG changes persist and active TEMPO groups temporarily overlay that prevailing state. This matches the replay-configuration comparison method.",
        ])
    for month in months:
        for mode, label in (("legacy_compatibility", "Legacy compatibility"), ("quality_gated", "Quality gated")):
            summary = month["monthly_summary"][mode]
            lines.append(
                "| {period} | {taf_count} | {label} | {wind_direction} | {wind_speed} | {wind_gust} | {visibility} | {rain_occurrence} | {cloud_amount} | {cloud_base} |".format(
                    period=month["period"],
                    taf_count=f"{month['scored_taf_count']}/{month['source_taf_count']}",
                    label=label,
                    **{metric: _format_score(summary[metric]["legacy_score_percent"]) for metric in METRICS},
                )
            )
    lines.extend([
        "",
        "Scores are an equal average of each scored issuance, matching the original monthly recap aggregation. The hourly denominators remain in the machine-readable JSON and issuance CSV.",
        "",
        "## Reconciliation With Stored Workbook Recap",
        "",
        "The source recap values below are cached workbook results. They are compared only where the cached cell is numeric; `#REF!` and `#VALUE!` are reported as unavailable rather than converted to a result.",
        "",
        "| Period | Metric | Stored recap | Independent legacy | Difference (pp) |",
        "| --- | --- | ---: | ---: | ---: |",
    ])
    for month in months:
        cached = month["source_recap_cached_values"]
        independent = month["monthly_summary"]["legacy_compatibility"]
        for metric in METRICS:
            value = cached.get(metric)
            try:
                stored = float(value)
            except (TypeError, ValueError):
                lines.append(
                    f"| {month['period']} | {metric.replace('_', ' ')} | {value or 'unavailable'} | "
                    f"{_format_score(independent[metric]['legacy_score_percent'])} | unavailable |"
                )
                continue
            calculated = independent[metric]["legacy_score_percent"]
            difference = stored - calculated if calculated is not None else None
            lines.append(
                f"| {month['period']} | {metric.replace('_', ' ')} | {stored:.2f}% | "
                f"{_format_score(calculated)} | {difference:+.2f} |"
            )
    lines.extend(["", "## Source and Quality Gates", ""])
    for month in months:
        quality = month["quality"]
        issues = []
        if quality["rain_timestamp_normalized_count"]:
            issues.append(f"rain timestamp normalized from matched METAR hour: {quality['rain_timestamp_normalized_count']} rows")
        if quality["rain_timestamp_unaligned_count"]:
            issues.append(f"unaligned rain timestamp remains unscored: {quality['rain_timestamp_unaligned_count']} rows")
        if quality["invalid_calendar_count"]:
            issues.append(f"invalid calendar timestamp: {quality['invalid_calendar_count']} rows")
        if quality["non_hourly_count"]:
            issues.append(f"non-hourly METAR: {quality['non_hourly_count']} rows")
        if month["repaired_validity_rows"]:
            source_rows = ", ".join(str(row) for row in month["repaired_validity_rows"])
            issues.append(f"impossible TAF validity repaired to 24 h: source row(s) {source_rows}")
        lines.append(f"- **{month['period']}**: " + ("; ".join(issues) if issues else "no canonical timestamp gate triggered.") )
    lines.extend([
        "",
        "## Interpretation",
        "",
        "- The legacy result is a reproducibility reference, not the operational skill score: its TEMPO rule can award full credit for fewer than half of temporary-hour hits.",
        "- March and May rain timestamps are normalized only because their HUJAN day/hour sequence matches the paired METAR sequence. Every affected hourly row is labelled `timestamp_normalized_from_row_alignment` in the audit CSV.",
        "- Three impossible TAF validity windows are evaluated as 24-hour windows under an explicit repair rule. Their original encoded end time and duration remain in the issuance and hourly audit files.",
        "- Visibility missing in the original legacy formula cache is treated as high visibility only in the legacy view. The quality-gated view marks it unavailable instead.",
        "- Cloud scores are parser-derived in both views; the original cloud formulas include incompatible spreadsheet functions and must not be used as an exact calculation source.",
        "",
        "## Audit Files",
        "",
        "- `historical_preconfigured_taf_hourly.csv`: one fully auditable row per TAF validity hour.",
        "- `historical_preconfigured_taf_issuance.csv`: one result per source TAF issuance and category.",
        "- `historical_preconfigured_taf_monthly.json`: machine-readable monthly result and source-quality summary.",
    ])
    report = output_dir / "HISTORICAL_PRECONFIGURED_TAF_RECALCULATION.md"
    report.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return report


def run_recalculation(
    downloads_dir: Path,
    metar_dir: Path,
    output_dir: Path,
    *,
    structured_taf_state: bool = False,
) -> dict[str, Any]:
    """Run the complete January-May recalculation and export its audit trail."""
    output_dir.mkdir(parents=True, exist_ok=True)
    all_hourly: list[dict[str, Any]] = []
    all_issuance: list[dict[str, Any]] = []
    months: list[dict[str, Any]] = []
    for period, filename in WORKBOOKS:
        workbook_path = downloads_dir / filename
        metar_path = metar_dir / f"metar_wawp_{period}.csv"
        if not workbook_path.exists():
            raise FileNotFoundError(f"Source workbook not found: {workbook_path}")
        if not metar_path.exists():
            raise FileNotFoundError(f"Canonical METAR file not found: {metar_path}")
        hourly, issuance, report = verify_month(
            period,
            workbook_path,
            metar_path,
            structured_taf_state=structured_taf_state,
        )
        rows = _read_metar_rows(metar_path)
        quality = {
            "rain_timestamp_normalized_count": sum(
                row.get("observed_at") is not None
                and not _is_true(row.get("rain_source_timestamp_matches_observed_at"))
                and _is_true(row.get("metar_time_group_matches_rain_timestamp"))
                for row in rows
            ),
            "rain_timestamp_unaligned_count": sum(
                row.get("observed_at") is not None
                and not _is_true(row.get("rain_source_timestamp_matches_observed_at"))
                and not _is_true(row.get("metar_time_group_matches_rain_timestamp"))
                for row in rows
            ),
            "invalid_calendar_count": sum(
                row.get("observed_at_status") in INVALID_OBSERVATION_STATUSES for row in rows
            ),
            "non_hourly_count": sum(row.get("metar_minute") not in {0, None} for row in rows),
        }
        months.append({
            **report,
            "quality": quality,
            "monthly_summary": {
                "legacy_compatibility": _monthly_aggregate(issuance, "legacy_compatibility"),
                "quality_gated": _monthly_aggregate(issuance, "quality_gated"),
            },
        })
        all_hourly.extend(hourly)
        all_issuance.extend(issuance)
    _write_csv(output_dir / "historical_preconfigured_taf_hourly.csv", all_hourly)
    _write_csv(output_dir / "historical_preconfigured_taf_issuance.csv", all_issuance)
    payload = {
        "generated_at_utc": datetime.now(UTC).isoformat().replace("+00:00", "Z"),
        "source_column": "TAF!C",
        "taf_group_state_mode": "structured" if structured_taf_state else "legacy_workbook_compatibility",
        "modes": ["legacy_compatibility", "quality_gated"],
        "months": months,
    }
    (output_dir / "historical_preconfigured_taf_monthly.json").write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")
    report_path = build_report(months, output_dir, structured_taf_state=structured_taf_state)
    return {"output_dir": str(output_dir), "report": str(report_path), "months": months}


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--downloads-dir", type=Path, default=Path.home() / "Downloads")
    parser.add_argument(
        "--metar-dir",
        type=Path,
        default=Path(r"D:\UJI_PERFORMA_MODEL\VERIFICATION_REPORTS\metar_standalone\canonical"),
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path(r"D:\UJI_PERFORMA_MODEL\VERIFICATION_REPORTS\historical_preconfigured_taf_recalculation"),
    )
    parser.add_argument(
        "--structured-taf-state",
        action="store_true",
        help="Evaluate BECMG/TEMPO using structured TAF state for a fair replay comparison.",
    )
    args = parser.parse_args()
    result = run_recalculation(
        args.downloads_dir,
        args.metar_dir,
        args.output_dir,
        structured_taf_state=args.structured_taf_state,
    )
    print(json.dumps({"output_dir": result["output_dir"], "report": result["report"]}, indent=2))


if __name__ == "__main__":
    main()
